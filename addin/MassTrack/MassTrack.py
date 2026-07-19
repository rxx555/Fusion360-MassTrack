"""MassTrack Fusion 360 add-in for opt-in CAD mass export.

Marked components/bodies export to a per-document CSV in a chosen output
folder; each document gets a folder named after it, holding mass.csv.
Volume cm^3, mass g,
density kg/m^3 (derived mass/volume, independent of Fusion's stored density).
"""

import adsk.core
import adsk.fusion
import traceback
import os
import sys
import csv
import json
import re
import uuid
import shutil
import subprocess
import datetime
import math

ADDIN_DIR = os.path.dirname(os.path.realpath(__file__))
DATA_DIR = os.path.join(os.path.expanduser("~"), ".masstrack")
try:
    os.makedirs(DATA_DIR, exist_ok=True)
except Exception:
    pass
SETTINGS_PATH = os.path.join(DATA_DIR, "settings.json")
LOG_PATH = os.path.join(DATA_DIR, "masstrack.log")

ATTR_GROUP = "MassTrack"
ATTR_NAME = "include"
ATTR_OVERRIDE = "massOverrideG"     # grams per unit, stored as string
ATTR_NOTE = "massNote"              # spec / measured

CSV_HEADER = ["key", "kind", "document", "name", "path", "material", "qty",
              "volume_cm3_each", "cad_mass_g_each", "mass_source",
              "mass_g_each", "mass_g_total", "density_kgm3",
              "flag", "note", "timestamp", "subassembly", "file_url"]

CMDS = [
    ("mtxGenerate", "MassTrack: Generate",
     "Export this document's marked parts to its CSV now."),
    ("mtxOpenXlsx", "MassTrack: Open in Excel",
     "Generate, then open this file's CSV in Excel."),
    ("mtxDiagram", "MassTrack: Diagram",
     "Capture the current 3D view into a Diagram sheet in the workbook, "
     "with a numbered parts legend. Set the view up first; there is no "
     "API to auto-explode."),
    ("mtxSnapshot", "MassTrack: Snapshot",
     "Force a labelled point onto the mass-history curve (e.g. 'isogrid v2')."),
    ("mtxMark", "MassTrack: Mark",
     "Include the selected components/bodies in the export (uses CAD mass)."),
    ("mtxSetKnown", "MassTrack: Set Mass",
     "Assign a measured mass (g). Overrides CAD mass. "
     "For electronics and parts whose CAD mass is meaningless."),
    ("mtxClearKnown", "MassTrack: Clear Mass",
     "Revert the selection to CAD-derived mass (keeps it marked)."),
    ("mtxUnmark", "MassTrack: Unmark",
     "Remove the selected components/bodies from the export."),
    ("mtxShow", "MassTrack: Show Marked",
     "List everything currently marked in this document."),
    ("mtxHighlight", "MassTrack: Highlight Marked",
     "Select all marked parts so they light up in the browser tree and canvas."),
    ("mtxSetFolder", "MassTrack: Output Folder",
     "Choose the folder where the CSV files are written."),
    ("mtxSetHub", "MassTrack: Set Hub URL",
     "Paste your Fusion Team base URL once (e.g. https://xxxx.autodesk360.com) "
     "so each row can link to the file. Access still needs an authorised login."),
]

_app = None
_ui = None
_handlers = []
_controls = []
_cmd_defs = []
_created = []       # (cmd_def, commandCreated handler) pairs wired this session
_doc_handler = None  # the documentSaved handler


def _log(msg):
    try:
        with open(LOG_PATH, "a") as f:
            f.write("%s  %s\n" % (
                datetime.datetime.now().isoformat(timespec="seconds"), msg))
    except Exception:
        pass


def _load_outdir():
    try:
        with open(SETTINGS_PATH) as f:
            return json.load(f).get("outdir") or None
    except Exception:
        return None


def _save_outdir(path):
    _save_setting("outdir", path)


def _prompt_outdir():
    dlg = _ui.createFolderDialog()
    dlg.title = "MassTrack output folder"
    if dlg.showDialog() == adsk.core.DialogResults.DialogOK:
        _save_outdir(dlg.folder)
        return dlg.folder
    return None


def _ensure_outdir():
    outdir = _load_outdir()
    if not outdir or not os.path.isdir(outdir):
        outdir = _prompt_outdir()
    return outdir


def _load_setting(field):
    try:
        with open(SETTINGS_PATH) as f:
            return json.load(f).get(field) or ""
    except Exception:
        return ""


def _save_setting(field, value):
    try:
        data = {}
        if os.path.isfile(SETTINGS_PATH):
            with open(SETTINGS_PATH) as f:
                data = json.load(f)
        data[field] = value
        with open(SETTINGS_PATH, "w") as f:
            json.dump(data, f)
    except Exception:
        _log("save setting failed:\n" + traceback.format_exc())


def _url_for_datafile(df):
    """Fusion Team web URL for a given DataFile, or '' if unavailable.

    Format (verified against a live hub):
      <hub>/g/projects/<numeric project id>/data/<b64url folder urn>/<b64url lineage urn>/overview
    Hub subdomain comes from the one-time 'Set Hub URL' setting since the API
    doesn't expose it cleanly.
    """
    if not df:
        return ""
    try:
        lineage = df.id                        # urn:adsk.wipprod:dm.lineage:...
        folder = df.parentFolder.id if df.parentFolder else ""
        projraw = df.parentProject.id if df.parentProject else ""
    except Exception:
        return ""
    base = _load_setting("hub_base")
    if not (base and lineage and folder and projraw):
        return ""
    import base64
    # parentProject.id is 'a.<base64(business:<hub>#<projectnum>)>'; the web URL
    # uses the numeric project id that follows '#'
    proj = projraw
    try:
        if projraw.startswith("a."):
            dec = base64.b64decode(projraw[2:] + "===").decode("utf-8", "replace")
            proj = dec.rsplit("#", 1)[-1]
    except Exception:
        pass

    def _b64(u):
        return base64.urlsafe_b64encode(u.encode()).decode().rstrip("=")

    return "%s/g/projects/%s/data/%s/%s/overview" % (
        base.rstrip("/"), proj, _b64(folder), _b64(lineage))


def _doc_link_url(doc_dir):
    """Fusion Team web URL for the active (master) document, or '' if local."""
    try:
        return _url_for_datafile(_app.activeDocument.dataFile)
    except Exception:
        return ""


def _component_source_datafile(comp, design):
    """DataFile of the external file a component comes from (its own sub-assembly
    / component file inserted as a reference), or None when the component is
    internal to the active document and so has no separate file to open."""
    try:
        occs = design.rootComponent.allOccurrencesByComponent(comp)
        if occs.count == 0:
            return None
        occ = occs.item(0)
        if not occ.isReferencedComponent:
            return None
        src = occ.component.parentDesign.parentDocument.dataFile
        active = _app.activeDocument.dataFile
        if src and (not active or src.id != active.id):
            _log("xref link: %s -> %s" % (comp.name, src.name))
            return src
    except Exception:
        _log("xref source lookup failed for %s:\n%s"
             % (getattr(comp, "name", "?"), traceback.format_exc(limit=2)))
    return None


def _entity_link_url(ent, kind, design, master_url):
    """URL that opens the file the marked entity lives in: its own component /
    sub-assembly file when that is an external reference, else the master."""
    try:
        comp = ent if kind == "component" else ent.parentComponent
        df = _component_source_datafile(comp, design)
        if df:
            u = _url_for_datafile(df)
            if u:
                return u
    except Exception:
        _log("entity link failed:\n" + traceback.format_exc())
    return master_url


def _open_externally(path, background=False):
    """Open a file in its default app (Excel for .csv on most setups). With
    background=True the file opens without bringing Excel in front of Fusion
    (macOS 'open -g'), so a silent refresh does not steal the user's focus."""
    try:
        if sys.platform == "darwin":
            base = ["open", "-g"] if background else ["open"]
            r = subprocess.run(base + ["-a", "Microsoft Excel", path])
            if r.returncode != 0:
                subprocess.run(base + [path])
        elif sys.platform.startswith("win"):
            os.startfile(path)
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        try:
            if sys.platform == "darwin":
                subprocess.Popen((["open", "-g"] if background else ["open"])
                                 + [path])
            else:
                os.startfile(path)
        except Exception:
            _log("open failed:\n" + traceback.format_exc())


_automation_warned = False


def _warn_automation_once():
    """One-time nudge when macOS blocks Excel automation (TCC not granted)."""
    global _automation_warned
    if _automation_warned:
        return
    _automation_warned = True
    try:
        _ui.messageBox(
            "MassTrack could not tell Excel to refresh the open workbook.\n\n"
            "Enable it in System Settings > Privacy & Security > Automation > "
            "Autodesk Fusion > Microsoft Excel, then regenerate.\n\n"
            "Until then, close the workbook in Excel before regenerating so the "
            "new data can load.", "MassTrack")
    except Exception:
        pass


def _excel_save_close(path):
    """macOS: if Excel already has this workbook open, save it (keeping the
    user's Work-sheet edits) and close it, so the rebuild that follows is not
    masked by Excel's cached copy. Excel will not hot-reload a file changed on
    disk, so a snapshot/regenerate otherwise looks like it did nothing until the
    user manually closes and reopens. No-op if Excel isn't running or the
    workbook isn't open. Windows locks the file instead, handled at the call
    site.

    Returns True when Excel was driven cleanly (including the nothing-to-do
    case), False when the workbook could not be saved/closed (e.g. macOS
    Automation permission denied, error -1743) so the caller can surface it.
    """
    if sys.platform != "darwin":
        return True
    name = os.path.basename(path)
    # Address the workbook explicitly by name. 'save'/'close' on a loop variable
    # from 'repeat with w in workbooks' fails with a -50 parameter error, which
    # left the stale copy open and made every refresh look like a no-op.
    script = (
        'on run argv\n'
        '  set n to item 1 of argv\n'
        '  if application "Microsoft Excel" is running then\n'
        '    tell application "Microsoft Excel"\n'
        '      if (exists workbook n) then\n'
        '        save workbook n\n'
        '        close workbook n saving yes\n'
        '      end if\n'
        '    end tell\n'
        '  end if\n'
        'end run\n')
    try:
        r = subprocess.run(["osascript", "-e", script, name],
                           capture_output=True, timeout=25)
        if r.returncode != 0:
            err = (r.stderr or b"").decode("utf-8", "replace").strip()
            _log("excel save/close osascript rc=%d: %s" % (r.returncode, err))
            return False
        return True
    except Exception:
        _log("excel save/close failed:\n" + traceback.format_exc())
        return False


def _excel_active_sheet(path):
    """Name of the active sheet in the open workbook, or None. Lets a refresh
    reopen on the sheet the user was viewing instead of jumping to Overview."""
    if sys.platform != "darwin":
        return None
    name = os.path.basename(path)
    script = (
        'on run argv\n'
        '  set n to item 1 of argv\n'
        '  if application "Microsoft Excel" is running then\n'
        '    tell application "Microsoft Excel"\n'
        '      if (exists workbook n) then\n'
        '        return name of active sheet of workbook n\n'
        '      end if\n'
        '    end tell\n'
        '  end if\n'
        '  return ""\n'
        'end run\n')
    try:
        r = subprocess.run(["osascript", "-e", script, name],
                           capture_output=True, timeout=25)
        if r.returncode == 0:
            s = (r.stdout or b"").decode("utf-8", "replace").strip()
            return s or None
    except Exception:
        _log("excel active-sheet read failed:\n" + traceback.format_exc())
    return None


def _refresh_workbook(doc_dir):
    """Rebuild the workbook and show it, refreshing Excel even when the file is
    already open (close it first, rebuild, reopen). Returns the path or None."""
    out = os.path.join(doc_dir, os.path.basename(doc_dir) + ".xlsx")
    active = _excel_active_sheet(out)          # remember the user's sheet
    if not _excel_save_close(out):
        _warn_automation_once()
    built = _build_workbook(doc_dir, active_sheet=active)
    if built:
        _open_externally(built, background=True)   # refresh without stealing focus
    return built


# "Diagram" is deliberately absent: _add_diagram_sheet owns it, and a normal
# regenerate must preserve it (only these sheets are deleted and rebuilt).
GENERATED_SHEETS = ("Overview", "Parts", "Simple", "History", "Trend")


def _num(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return s
    t = str(s).strip()
    if t == "":
        return None
    low = t.lower()
    if low in ("nan", "inf", "-inf", "+inf"):
        return None
    try:
        f = float(t)
    except (ValueError, TypeError):
        return s
    if f.is_integer() and ("." not in t and "e" not in low):
        try:
            return int(t)
        except (ValueError, TypeError):
            return f
    return f


def _read_csv(path):
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.reader(f))
    except Exception:
        return []


def _build_workbook(doc_dir, active_sheet=None):
    try:
        vendor = os.path.join(ADDIN_DIR, "vendor")
        if vendor not in sys.path:
            sys.path.insert(0, vendor)
        import datetime
        import openpyxl  # noqa: F401
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.layout import Layout, ManualLayout

        FN = "Calibri"
        INK, MUTE = "1F2933", "5B6670"
        ACCENT, ACCENTD, ACCENTL = "E06A1B", "A8480F", "FCE9D6"
        HDRTX = "FFFFFF"
        BAND, GRID = "F8F5F1", "E5DCD3"
        AMBER, MANUAL, TOTBG = "F6D488", "DCE6F1", "F6EEE6"

        def _fill(c):
            return PatternFill("solid", fgColor=c)

        RIGHT = Alignment(horizontal="right")
        LEFT = Alignment(horizontal="left")
        VCEN = Alignment(vertical="center")
        rule_hdr = Border(bottom=Side(style="medium", color=ACCENTD))
        grp_rule = Border(top=Side(style="thin", color=ACCENT))
        tot_rule = Border(top=Side(style="medium", color=ACCENTD))
        card_edge = Border(left=Side(style="thick", color=ACCENT))
        F_MASS, F_INT = '#,##0.0', '#,##0'
        F_KG = '#,##0.000" kg"'
        F_G = '#,##0" g"'

        def _hdr(ws, row, labels, widths, aligns=None):
            for j, lab in enumerate(labels, start=1):
                c = ws.cell(row, j, lab)
                c.font = Font(name=FN, bold=True, color=HDRTX, size=11)
                c.fill = _fill(ACCENT)
                c.border = rule_hdr
                c.alignment = VCEN if not aligns else Alignment(
                    horizontal=aligns[j - 1], vertical="center")
            for j, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(j)].width = w
            ws.row_dimensions[row].height = 22

        def _band(ws, row, ncol):
            for c in range(1, ncol + 1):
                ws.cell(row, c).fill = _fill(BAND)

        def _link(cell, url):
            if str(url).strip():
                cell.hyperlink = url
                try:
                    cell.style = "Hyperlink"
                except Exception:
                    cell.font = Font(name=FN, color=ACCENT, underline="single")

        def _autofit(wsx, cap=55, floor=6, pad=2):
            merged = set()
            for rng in wsx.merged_cells.ranges:
                for r_ in wsx[str(rng)]:
                    for c_ in r_:
                        merged.add(c_.coordinate)
            wide = {}
            for r_ in wsx.iter_rows():
                for c_ in r_:
                    if c_.value is None or c_.coordinate in merged:
                        continue
                    try:
                        col = c_.column_letter
                    except Exception:
                        continue
                    ln = len(str(c_.value))
                    if ln > wide.get(col, 0):
                        wide[col] = ln
            for col, ln in wide.items():
                wsx.column_dimensions[col].width = min(cap, max(floor, ln + pad))

        def _mass_chart(src, effcol, catcol, lastrow, title):
            c = LineChart()
            c.title = title
            # No axis TITLES: in Excel, openpyxl axis titles render on top of the
            # tick numbers and cannot be positioned deterministically. Units live
            # in the chart title instead; the numbers themselves are self-evident.
            c.y_axis.delete = False
            c.x_axis.delete = False
            c.y_axis.numFmt = '#,##0'
            c.y_axis.majorTickMark = "out"
            c.x_axis.majorTickMark = "out"
            c.height, c.width = 8, 15
            c.legend = None
            # small margins so the numbers have room and the value axis sits in
            # off the chart edge (no axis titles to accommodate now)
            c.plot_area.layout = Layout(manualLayout=ManualLayout(
                xMode="edge", yMode="edge", x=0.10, y=0.16, w=0.86, h=0.66))
            c.add_data(Reference(src, min_col=effcol, max_col=effcol,
                                 min_row=1, max_row=lastrow), titles_from_data=True)
            c.set_categories(Reference(src, min_col=catcol, min_row=2, max_row=lastrow))
            try:
                s = c.series[0]
                s.smooth = True
                s.graphicalProperties.line.solidFill = ACCENT
                s.graphicalProperties.line.width = 22000
            except Exception:
                pass
            return c

        docname = os.path.basename(doc_dir)
        stamp = datetime.date.today().isoformat()
        mass = _read_csv(os.path.join(doc_dir, "mass.csv"))
        totals = _read_csv(os.path.join(doc_dir, "totals_history.csv"))
        parts = _read_csv(os.path.join(doc_dir, "parts_history.csv"))
        mrows = mass[1:] if len(mass) > 1 else []

        recs = []
        for r in mrows:
            if len(r) < 18:
                r = r + [""] * (18 - len(r))
            qty = _num(r[6]) or 0
            cad_each = _num(r[8]) or 0
            eff_tot = _num(r[11]) or 0
            try:
                cad_tot = float(cad_each) * float(qty) if cad_each else 0.0
            except (TypeError, ValueError):
                cad_tot = 0.0
            recs.append({"key": r[0], "name": r[3], "sub": r[16] or "(unassigned)",
                         "material": r[5], "qty": qty, "cad_tot": cad_tot,
                         "source": r[9], "eff_tot": eff_tot, "flag": r[13],
                         "note": r[14], "url": r[17]})

        recs.sort(key=lambda x: (str(x["sub"]).lower(), str(x["name"]).lower()))
        total_eff = sum((x["eff_tot"] or 0) for x in recs)
        n_parts = len(recs)
        n_flag = sum(1 for x in recs if str(x["flag"]).strip())
        sub_map = {}
        for x in recs:
            m, c = sub_map.get(x["sub"], (0.0, 0))
            sub_map[x["sub"]] = (m + (x["eff_tot"] or 0), c + 1)
        sub_rows = sorted(sub_map.items(), key=lambda kv: kv[1][0], reverse=True)
        mat_map = {}
        for x in recs:
            mname = str(x["material"]).strip() or "(no material)"
            m, c = mat_map.get(mname, (0.0, 0))
            mat_map[mname] = (m + (x["eff_tot"] or 0), c + 1)
        mat_rows = sorted(mat_map.items(), key=lambda kv: kv[1][0], reverse=True)

        out = os.path.join(doc_dir, docname + ".xlsx")
        # keep user sheets, rewrite only the generated ones
        if os.path.isfile(out):
            try:
                wb = load_workbook(out)
                for nm in GENERATED_SHEETS:
                    if nm in wb.sheetnames:
                        del wb[nm]
            except Exception:
                bak = os.path.join(doc_dir, docname + ".backup.xlsx")
                try:
                    shutil.copyfile(out, bak)
                    _log("workbook load failed, backed up to " + bak)
                except Exception:
                    _log("workbook load failed, backup failed:\n"
                         + traceback.format_exc())
                wb = Workbook()
                wb.remove(wb.active)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        ws = wb.create_sheet("Overview")
        ws.sheet_view.showGridLines = True
        for col, w in (("A", 30), ("B", 15), ("C", 13), ("D", 3)):
            ws.column_dimensions[col].width = w

        ws.merge_cells("A1:D1")
        ws["A1"] = docname
        ws["A1"].font = Font(name=FN, size=22, bold=True, color=ACCENTD)
        ws.merge_cells("A2:D2")
        ws["A2"] = stamp
        ws["A2"].font = Font(name=FN, size=10, color=MUTE)

        # KPI card
        for rr in range(4, 7):
            for cc in range(1, 4):
                ws.cell(rr, cc).fill = _fill(ACCENTL)
            ws.cell(rr, 1).border = card_edge
        ws.merge_cells("A4:C4")
        ws["A4"] = "TOTAL EFFECTIVE MASS"
        ws["A4"].font = Font(name=FN, size=9, bold=True, color=MUTE)
        ws.merge_cells("A5:B6")
        ws["A5"] = total_eff / 1000.0
        ws["A5"].number_format = F_KG
        ws["A5"].font = Font(name=FN, size=30, bold=True, color=ACCENT)
        ws["A5"].alignment = Alignment(vertical="center")
        ws["C5"] = total_eff
        ws["C5"].number_format = F_G
        ws["C5"].font = Font(name=FN, size=11, color=MUTE)
        ws["C5"].alignment = RIGHT
        ws["C6"] = str(n_parts) + (" part" if n_parts == 1 else " parts")
        ws["C6"].font = Font(name=FN, size=10, color=MUTE)
        ws["C6"].alignment = RIGHT

        ws["A8"] = "Marked parts"
        ws["A8"].font = Font(name=FN, bold=True, color=INK)
        ws["B8"] = n_parts
        ws["B8"].number_format = F_INT
        ws["B8"].alignment = RIGHT
        ws["A9"] = "Flagged parts"
        ws["A9"].font = Font(name=FN, bold=True, color=INK)
        ws["B9"] = n_flag
        ws["B9"].number_format = F_INT
        ws["B9"].alignment = RIGHT

        sr = 12
        ws.cell(sr - 1, 1, "By subassembly").font = Font(
            name=FN, size=12, bold=True, color=ACCENTD)
        _hdr(ws, sr, ("Subassembly", "Mass (g)", "Parts"), (30, 15, 13),
             aligns=("left", "right", "right"))
        for i, (name, (m, c)) in enumerate(sub_rows):
            rr = sr + 1 + i
            if i % 2:
                _band(ws, rr, 3)
            ws.cell(rr, 1, name).font = Font(name=FN, color=INK)
            mc = ws.cell(rr, 2, m)
            mc.number_format = F_MASS
            mc.alignment = RIGHT
            pc = ws.cell(rr, 3, c)
            pc.number_format = F_INT
            pc.alignment = RIGHT
        trow = sr + 1 + len(sub_rows)
        for cc in range(1, 4):
            ws.cell(trow, cc).fill = _fill(TOTBG)
            ws.cell(trow, cc).border = tot_rule
        ws.cell(trow, 1, "TOTAL").font = Font(name=FN, bold=True, color=INK)
        tc = ws.cell(trow, 2, total_eff)
        tc.number_format = F_MASS
        tc.font = Font(name=FN, bold=True, color=INK)
        tc.alignment = RIGHT
        pc = ws.cell(trow, 3, n_parts)
        pc.font = Font(name=FN, bold=True, color=INK)
        pc.alignment = RIGHT
        ov_end = trow

        top = sorted(recs, key=lambda x: -(x["eff_tot"] or 0))[:5]
        if top:
            hb = trow + 2
            ws.cell(hb, 1, "Heaviest parts").font = Font(
                name=FN, size=12, bold=True, color=ACCENTD)
            for j, lab in enumerate(("Part", "Mass (g)", "Share")):
                hc = ws.cell(hb + 1, 1 + j, lab)
                hc.font = Font(name=FN, bold=True, color=HDRTX, size=11)
                hc.fill = _fill(ACCENT)
                hc.border = rule_hdr
                hc.alignment = Alignment(
                    horizontal=("left" if j == 0 else "right"), vertical="center")
            for i, x in enumerate(top):
                rr = hb + 2 + i
                if i % 2:
                    _band(ws, rr, 3)
                ws.cell(rr, 1, x["name"]).font = Font(name=FN, color=INK)
                mc = ws.cell(rr, 2, x["eff_tot"])
                mc.number_format = F_MASS
                mc.alignment = RIGHT
                sc = ws.cell(rr, 3, (x["eff_tot"] or 0) / total_eff if total_eff else 0)
                sc.number_format = '0.0%'
                sc.alignment = RIGHT
            ov_end = hb + 1 + len(top)

        if mat_rows:
            mb = ov_end + 2
            ws.cell(mb, 1, "By material").font = Font(
                name=FN, size=12, bold=True, color=ACCENTD)
            _hdr(ws, mb + 1, ("Material", "Mass (g)", "Parts"), (30, 15, 13),
                 aligns=("left", "right", "right"))
            for i, (name, (m, c)) in enumerate(mat_rows):
                rr = mb + 2 + i
                if i % 2:
                    _band(ws, rr, 3)
                ws.cell(rr, 1, name).font = Font(name=FN, color=INK)
                mc = ws.cell(rr, 2, m)
                mc.number_format = F_MASS
                mc.alignment = RIGHT
                pc = ws.cell(rr, 3, c)
                pc.number_format = F_INT
                pc.alignment = RIGHT
            mt = mb + 2 + len(mat_rows)
            for cc in range(1, 4):
                ws.cell(mt, cc).fill = _fill(TOTBG)
                ws.cell(mt, cc).border = tot_rule
            ws.cell(mt, 1, "TOTAL").font = Font(name=FN, bold=True, color=INK)
            tmc = ws.cell(mt, 2, total_eff)
            tmc.number_format = F_MASS
            tmc.font = Font(name=FN, bold=True, color=INK)
            tmc.alignment = RIGHT
            tpc = ws.cell(mt, 3, n_parts)
            tpc.font = Font(name=FN, bold=True, color=INK)
            tpc.alignment = RIGHT
            ov_end = mt

        wp = wb.create_sheet("Parts")
        wp.sheet_view.showGridLines = True
        headers = ["Name", "Subassembly", "Material", "Qty", "CAD mass (g)",
                   "Source", "Effective mass (g)", "Flag", "Note"]
        widths = [34, 20, 16, 8, 14, 10, 18, 14, 40]
        aligns = ["left", "left", "left", "right", "right", "left", "right",
                  "left", "left"]
        _hdr(wp, 1, headers, widths, aligns=aligns)
        wp.freeze_panes = "A2"
        orow, prev_sub = 2, None
        for x in recs:
            if prev_sub is not None and x["sub"] != prev_sub:
                orow += 1                       # gap between subassemblies
            prev_sub = x["sub"]
            nc = wp.cell(orow, 1, x["name"])
            _link(nc, x["url"])
            wp.cell(orow, 2, x["sub"]).font = Font(name=FN, color=INK)
            wp.cell(orow, 3, x["material"]).font = Font(name=FN, color=INK)
            qc = wp.cell(orow, 4, x["qty"])
            qc.number_format = F_INT
            qc.alignment = RIGHT
            cc = wp.cell(orow, 5, x["cad_tot"])
            cc.number_format = F_MASS
            cc.alignment = RIGHT
            wp.cell(orow, 6, x["source"]).font = Font(name=FN, color=MUTE)
            ec = wp.cell(orow, 7, x["eff_tot"])
            ec.number_format = F_MASS
            ec.font = Font(name=FN, bold=True, color=INK)
            ec.alignment = RIGHT
            fc = wp.cell(orow, 8, x["flag"])
            wp.cell(orow, 9, x["note"]).font = Font(name=FN, color=MUTE)
            if str(x["flag"]).strip():
                fc.fill = _fill(AMBER)
            if str(x["source"]).strip().lower() == "manual":
                ec.fill = _fill(MANUAL)
            orow += 1
        lastdata = orow - 1
        wp.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), max(1, lastdata))
        trow = orow + 1
        for c in range(1, len(headers) + 1):
            wp.cell(trow, c).fill = _fill(TOTBG)
            wp.cell(trow, c).border = tot_rule
        wp.cell(trow, 1, "TOTAL").font = Font(name=FN, bold=True, color=INK)
        tc = wp.cell(trow, 7, ("=SUM(G2:G%d)" % lastdata) if recs else 0)
        tc.number_format = F_MASS
        tc.font = Font(name=FN, bold=True, color=INK)
        tc.alignment = RIGHT

        wsx = wb.create_sheet("Simple")
        wsx.sheet_view.showGridLines = True
        _hdr(wsx, 1, ("Key", "Part name", "Mass (g)"), (16, 40, 14),
             aligns=("left", "left", "right"))
        wsx.freeze_panes = "A2"
        for i, x in enumerate(recs):
            rr = i + 2
            if i % 2:
                _band(wsx, rr, 3)
            wsx.cell(rr, 1, x["key"]).font = Font(name=FN, color=MUTE)
            nc = wsx.cell(rr, 2, x["name"])
            _link(nc, x["url"])
            mc = wsx.cell(rr, 3, x["eff_tot"])
            mc.number_format = F_MASS
            mc.alignment = RIGHT
        last = len(recs) + 1
        trow = last + 1
        for c in range(1, 4):
            wsx.cell(trow, c).fill = _fill(TOTBG)
            wsx.cell(trow, c).border = tot_rule
        wsx.cell(trow, 1, "TOTAL").font = Font(name=FN, bold=True, color=INK)
        tc = wsx.cell(trow, 3, "=SUM(C2:C%d)" % last if recs else 0)
        tc.number_format = F_MASS
        tc.font = Font(name=FN, bold=True, color=INK)
        tc.alignment = RIGHT

        trows = totals[1:] if len(totals) > 1 else []
        if trows:
            wh = wb.create_sheet("History")
            wh.sheet_view.showGridLines = True
            _hdr(wh, 1, ("No.", "Snapshot", "Label", "Parts", "CAD mass (g)",
                         "Effective mass (g)"), (6, 22, 22, 10, 16, 18),
                 aligns=("right", "left", "left", "right", "right", "right"))
            for i, row in enumerate(trows):
                rr = i + 2
                row = (row + [""] * 5)[:5]
                if i % 2:
                    _band(wh, rr, 6)
                n = wh.cell(rr, 1, i + 1)
                n.number_format = F_INT
                n.alignment = RIGHT
                n.font = Font(name=FN, color=MUTE)
                wh.cell(rr, 2, row[0]).font = Font(name=FN, color=INK)
                wh.cell(rr, 3, row[1]).font = Font(name=FN, color=INK)
                a = wh.cell(rr, 4, _num(row[2]))
                a.number_format = F_INT
                a.alignment = RIGHT
                b = wh.cell(rr, 5, _num(row[3]))
                b.number_format = F_MASS
                b.alignment = RIGHT
                d = wh.cell(rr, 6, _num(row[4]))
                d.number_format = F_MASS
                d.alignment = RIGHT
            hlast = len(trows) + 1
            if len(trows) >= 2:
                wh.add_chart(
                    _mass_chart(wh, 6, 1, hlast, "Total marked mass (g) per snapshot"),
                    "H2")
                ws.add_chart(
                    _mass_chart(wh, 6, 1, hlast, "Total marked mass (g) per snapshot"),
                    "A%d" % (ov_end + 3))

        ph = parts[1:] if len(parts) > 1 else []
        ts_order, keys_order, key_name, val = [], [], {}, {}
        for row in ph:
            if len(row) < 6:
                continue
            ts, key, name, m = row[0], row[2], row[3], row[5]
            if ts not in ts_order:
                ts_order.append(ts)
            if key not in keys_order:
                keys_order.append(key)
            key_name[key] = name
            val[(key, ts)] = m
        ts_order.sort()
        if keys_order and ts_order:
            wt = wb.create_sheet("Trend")
            wt.sheet_view.showGridLines = True
            header = ["No.", "Snapshot"] + [key_name[k] for k in keys_order]
            widths = [6, 20] + [max(12, min(28, len(str(key_name[k])) + 2))
                                for k in keys_order]
            _hdr(wt, 1, header, widths,
                 aligns=["right", "left"] + ["right"] * len(keys_order))
            wt.freeze_panes = "C2"
            cur, n_ts = {}, 0
            for ts in ts_order:
                for k in keys_order:
                    if (k, ts) in val:
                        cur[k] = _num(val[(k, ts)])
                rr = n_ts + 2
                if n_ts % 2:
                    _band(wt, rr, len(header))
                nn = wt.cell(rr, 1, n_ts + 1)
                nn.number_format = F_INT
                nn.alignment = RIGHT
                nn.font = Font(name=FN, color=MUTE)
                wt.cell(rr, 2, ts).font = Font(name=FN, color=INK)
                for j, k in enumerate(keys_order, start=3):
                    v = cur.get(k, None)
                    if v is not None:
                        c = wt.cell(rr, j, v)
                        c.number_format = F_MASS
                        c.alignment = RIGHT
                n_ts += 1
            if len(ts_order) >= 2:
                cp = LineChart()
                cp.title = "Mass per part (g) per snapshot"
                # no axis titles (they overlap the numbers in Excel); units are in
                # the chart title, numbers kept
                cp.y_axis.delete = False
                cp.x_axis.delete = False
                cp.y_axis.numFmt = '#,##0'
                cp.y_axis.majorTickMark = "out"
                cp.x_axis.majorTickMark = "out"
                cp.width = 24
                cp.height = max(10, 3 + 1.1 * len(keys_order))
                # push the plot in from the left edge; reserve the right band for
                # the legend and a bottom band for the x-axis numbers
                cp.plot_area.layout = Layout(manualLayout=ManualLayout(
                    xMode="edge", yMode="edge", x=0.07, y=0.12, w=0.74, h=0.74))
                cp.add_data(Reference(wt, min_col=3, max_col=2 + len(keys_order),
                                      min_row=1, max_row=1 + n_ts), titles_from_data=True)
                cp.set_categories(Reference(wt, min_col=1, min_row=2, max_row=1 + n_ts))
                for s in cp.series:
                    s.smooth = True
                if cp.legend is not None:
                    cp.legend.position = 'r'
                    cp.legend.overlay = False
                wt.add_chart(cp, "A" + str(n_ts + 4))

        # user-owned rows never overwritten; only the seeded header cells + the
        # sheet's own cosmetics (colour, gridlines) are re-stamped each rebuild
        # so a Work sheet made under an older palette picks up the current one.
        if "Work" not in wb.sheetnames:
            wk = wb.create_sheet("Work")
            wk.column_dimensions["A"].width = 24
            wk.column_dimensions["B"].width = 16
            wk["A1"] = "Workspace"
            wk["A3"] = "Current total (kg)"
            wk["B3"] = "=Overview!A5"
            wk["B3"].number_format = F_KG
            wk["B3"].alignment = RIGHT
        wk = wb["Work"]
        wk.sheet_view.showGridLines = True
        wk["A1"].font = Font(name=FN, size=16, bold=True, color=ACCENTD)
        wk["A3"].font = Font(name=FN, bold=True, color=INK)
        wk["B3"].font = Font(name=FN, bold=True, color=ACCENT)

        # Diagram: re-embed from the persistent PNG on EVERY rebuild. A
        # load->save cycle without Pillow drops embedded images, so the picture
        # must be re-added each time or it vanishes on the next regenerate.
        png = os.path.join(doc_dir, "diagram.png")
        if os.path.isfile(png):
            try:
                from openpyxl.drawing.image import Image as XLImage

                class _RawPNG(XLImage):
                    def __init__(self, p):
                        self.ref = p
                        self.format = "png"
                        self.anchor = "A1"
                        self.width, self.height = _png_size(p)

                    def _data(self):
                        with open(self.ref, "rb") as f:
                            return f.read()

                if "Diagram" in wb.sheetnames:
                    del wb["Diagram"]
                wd = wb.create_sheet("Diagram")
                wd.sheet_view.showGridLines = False
                wd["A1"] = docname + " (captured view)"
                wd["A1"].font = Font(name=FN, size=14, bold=True, color=ACCENTD)
                dimg = _RawPNG(png)
                if dimg.width > 900:
                    dimg.height = max(1, int(round(dimg.height * 900.0 / dimg.width)))
                    dimg.width = 900
                wd.add_image(dimg, "A3")
                for w_, letter in ((5, "N"), (34, "O"), (12, "P")):
                    wd.column_dimensions[letter].width = w_
                for j, lab in enumerate(("#", "Part", "Mass (g)")):
                    hc = wd.cell(2, 14 + j, lab)
                    hc.font = Font(name=FN, bold=True, color=HDRTX)
                    hc.fill = _fill(ACCENT)
                for i, x in enumerate(recs):
                    rr = 3 + i
                    nc = wd.cell(rr, 14, i + 1)
                    nc.alignment = RIGHT
                    wd.cell(rr, 15, x["name"]).font = Font(name=FN, color=INK)
                    mc = wd.cell(rr, 16, x["eff_tot"])
                    mc.number_format = F_MASS
                    mc.alignment = RIGHT
            except Exception:
                _log("diagram embed failed:\n" + traceback.format_exc())

        # generated sheets first, then user sheets
        for wsx in wb.worksheets:
            if wsx.title in GENERATED_SHEETS:
                _autofit(wsx)

        order = [n for n in GENERATED_SHEETS if n in wb.sheetnames]
        order += [n for n in wb.sheetnames if n not in order]
        wb._sheets.sort(key=lambda s: order.index(s.title))
        wb.active = 0
        if active_sheet and active_sheet in wb.sheetnames:
            try:
                wb.active = wb.sheetnames.index(active_sheet)
            except Exception:
                pass

        try:
            wb.properties.creator = ""
            wb.properties.lastModifiedBy = ""
            wb.properties.title = None
            wb.properties.company = None
        except Exception:
            pass

        tmp = out + ".tmp"
        wb.save(tmp)
        os.replace(tmp, out)
        return out
    except Exception:
        _log("workbook build failed:\n" + traceback.format_exc())
        return None


def _capture_view(path):
    """Save the active viewport as a PNG. True on success."""
    try:
        vp = _app.activeViewport
        try:
            opts = adsk.core.SaveImageFileOptions.create()
            opts.filename = path
            opts.width = 1600
            opts.height = 1000
            opts.isBackgroundTransparent = True
            if vp.saveAsImageFileWithOptions(opts) and os.path.isfile(path):
                return True
        except Exception:
            _log("saveAsImageFileWithOptions failed:\n"
                 + traceback.format_exc())
        if vp.saveAsImageFile(path, 1600, 1000) and os.path.isfile(path):
            return True
    except Exception:
        _log("view capture failed:\n" + traceback.format_exc())
    return False


def _png_size(path):
    with open(path, "rb") as f:
        h = f.read(24)
    if len(h) < 24 or h[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("not a PNG: " + path)
    return (int.from_bytes(h[16:20], "big"),
            int.from_bytes(h[20:24], "big"))


def _doc_dir(outdir, docname):
    """Per-document output subfolder, created if absent."""
    safe = re.sub(r"[^0-9A-Za-z._-]+", "_", docname).strip("_") or "untitled"
    d = os.path.join(outdir, safe)
    os.makedirs(d, exist_ok=True)
    return d


def _csv_path(outdir, docname):
    return os.path.join(_doc_dir(outdir, docname), "mass.csv")


TOTALS_HEADER = ["snapshot_ts", "label", "n_parts",
                 "total_cad_mass_g", "total_effective_mass_g"]
PARTS_HEADER = ["snapshot_ts", "label", "key", "name",
                "mass_source", "mass_g_total"]


def _append_csv(path, header, rows):
    exists = os.path.isfile(path)
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if not exists:
            w.writerow(header)
        w.writerows(rows)


def _load_last(path):
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {"total": None, "parts": {}}


def _save_last(path, data):
    try:
        with open(path, "w") as f:
            json.dump(data, f)
    except Exception:
        pass


def _entity_from_selection(ent):
    """Normalise a selection to the thing we stamp: Component or BRepBody."""
    occ = adsk.fusion.Occurrence.cast(ent)
    if occ:
        return occ.component
    body = adsk.fusion.BRepBody.cast(ent)
    if body:
        return body.nativeObject if body.nativeObject else body
    comp = adsk.fusion.Component.cast(ent)
    if comp:
        return comp
    return None


def _set_attr(target, name, value):
    a = target.attributes.itemByName(ATTR_GROUP, name)
    if a:
        a.value = value
    else:
        target.attributes.add(ATTR_GROUP, name, value)


def _ensure_included(target):
    if target.attributes.itemByName(ATTR_GROUP, ATTR_NAME) is None:
        target.attributes.add(ATTR_GROUP, ATTR_NAME, uuid.uuid4().hex[:8])


def _mark(entities):
    n = 0
    for ent in entities:
        target = _entity_from_selection(ent)
        if target is None:
            continue
        if target.attributes.itemByName(ATTR_GROUP, ATTR_NAME) is None:
            target.attributes.add(ATTR_GROUP, ATTR_NAME, uuid.uuid4().hex[:8])
            n += 1
    return n


def _set_known_mass(entities, grams, note):
    n = 0
    for ent in entities:
        target = _entity_from_selection(ent)
        if target is None:
            continue
        _ensure_included(target)
        _set_attr(target, ATTR_OVERRIDE, grams)
        _set_attr(target, ATTR_NOTE, note)
        n += 1
    return n


def _clear_known_mass(entities):
    n = 0
    for ent in entities:
        target = _entity_from_selection(ent)
        if target is None:
            continue
        for name in (ATTR_OVERRIDE, ATTR_NOTE):
            a = target.attributes.itemByName(ATTR_GROUP, name)
            if a:
                a.deleteMe()
                n += 1
    return n


def _unmark(entities):
    n = 0
    for ent in entities:
        target = _entity_from_selection(ent)
        if target is None:
            continue
        attr = target.attributes.itemByName(ATTR_GROUP, ATTR_NAME)
        if attr:
            attr.deleteMe()
            n += 1
    return n


def _marked_items(design):
    """Yield (key, kind, entity) for every marked component and body."""
    seen = set()
    for comp in design.allComponents:
        attr = comp.attributes.itemByName(ATTR_GROUP, ATTR_NAME)
        if attr and comp.entityToken not in seen:
            seen.add(comp.entityToken)
            yield attr.value, "component", comp
        for body in comp.bRepBodies:
            battr = body.attributes.itemByName(ATTR_GROUP, ATTR_NAME)
            if battr and body.entityToken not in seen:
                seen.add(body.entityToken)
                yield battr.value, "body", body


def _phys(ent, accuracy):
    try:
        return ent.getPhysicalProperties(accuracy)
    except Exception:
        return ent.physicalProperties


def _descendant_tokens(comp, root):
    """entityTokens of every component nested below comp.

    Components aren't hashable, so identity goes through entityToken. Used by
    the double-count guard since mass properties are recursive.
    """
    result = set()
    try:
        for occ in root.allOccurrencesByComponent(comp):
            stack = list(occ.childOccurrences)
            while stack:
                child = stack.pop()
                try:
                    result.add(child.component.entityToken)
                except Exception:
                    pass
                stack.extend(list(child.childOccurrences))
    except Exception:
        pass
    return result


def _top_subassembly(comp, root):
    """Name of the top-level sub-assembly this component sits under."""
    try:
        if comp == root:
            return "(root)"
        occs = root.allOccurrencesByComponent(comp)
        if occs and occs.count > 0:
            first = occs.item(0).fullPathName.split("+")[0]
            return first.rsplit(":", 1)[0] if ":" in first else first
    except Exception:
        pass
    return comp.name


def _props_row(key, kind, ent, design, ts, docname):
    root = design.rootComponent
    accuracy = adsk.fusion.CalculationAccuracy.HighCalculationAccuracy
    flags = []

    if kind == "component":
        comp = ent
        qty = root.allOccurrencesByComponent(comp).count or 1
        props = _phys(comp, accuracy)
        mats = set(b.material.name if b.material else "(none)"
                   for b in comp.bRepBodies)
        if not mats:
            material = "(assembly)"
        elif len(mats) == 1:
            material = mats.pop()
        else:
            material = "mixed"
        name = comp.name
        path = comp.name
        if comp.occurrences.count > 0:
            flags.append("assembly:recursive-mass")
    else:  # body
        body = ent
        comp = body.parentComponent
        qty = root.allOccurrencesByComponent(comp).count or 1
        props = _phys(body, accuracy)
        material = body.material.name if body.material else "(none)"
        name = body.name
        path = "%s/%s" % (comp.name, body.name)

    vol_cm3 = props.volume
    cad_mass_g = props.mass * 1000.0
    dens = (props.mass / (vol_cm3 * 1e-6)) if vol_cm3 > 0 else 0.0

    # known-mass override beats CAD when set
    ov = ent.attributes.itemByName(ATTR_GROUP, ATTR_OVERRIDE)
    note_a = ent.attributes.itemByName(ATTR_GROUP, ATTR_NOTE)
    note = note_a.value if note_a else ""
    eff_mass_g = cad_mass_g
    source = "cad"
    if ov and str(ov.value).strip():
        try:
            eff_mass_g = float(ov.value)
            source = "manual"
        except ValueError:
            flags.append("bad-override:%s" % ov.value)

    if source == "cad" and material.strip() == "Steel":   # untouched default
        flags.append("default-material?")
    if vol_cm3 <= 1e-4 and source == "cad":
        flags.append("zero-volume")

    subassembly = _top_subassembly(comp, root)

    return [key, kind, docname, name, path, material, str(qty),
            "%.4f" % vol_cm3, "%.3f" % cad_mass_g, source,
            "%.3f" % eff_mass_g, "%.3f" % (eff_mass_g * qty),
            "%.1f" % dens, ";".join(flags), note, ts, subassembly]


def _update_history(outdir, docname, ts, rows, label, force):
    """Append totals + per-part history on change; force writes a labelled point."""
    # CSV_HEADER indices: 0 key, 3 name, 6 qty, 8 cad_each, 9 source, 11 eff_total
    total_eff = sum(float(r[11]) for r in rows)
    total_cad = sum(float(r[8]) * int(r[6]) for r in rows)
    cur_parts = {r[0]: "%.3f" % float(r[11]) for r in rows}

    d = _doc_dir(outdir, docname)
    last_path = os.path.join(d, "last.json")
    last = _load_last(last_path)

    changed = force or last["total"] is None \
        or abs(total_eff - (last["total"] or 0.0)) > 1e-3 \
        or cur_parts != last.get("parts", {})
    if not changed:
        return False, total_eff

    _append_csv(os.path.join(d, "totals_history.csv"), TOTALS_HEADER,
                [[ts, label, str(len(rows)),
                  "%.3f" % total_cad, "%.3f" % total_eff]])

    changed_rows = [[ts, label, r[0], r[3], r[9], r[11]] for r in rows
                    if force or cur_parts[r[0]] != last.get("parts", {}).get(r[0])]
    if changed_rows:
        _append_csv(os.path.join(d, "parts_history.csv"), PARTS_HEADER,
                    changed_rows)

    _save_last(last_path, {"total": total_eff, "parts": cur_parts})
    return True, total_eff


def _doc_is_saved():
    """True only if the active document has been saved at least once."""
    try:
        return _app.activeDocument.dataFile is not None
    except Exception:
        return False


def _doc_name():
    # versionless name so the folder is stable across saves (dataFile.name has
    # no " v12" suffix; strip it as a fallback for the local/edge cases)
    try:
        df = _app.activeDocument.dataFile
        if df and df.name:
            return df.name
    except Exception:
        pass
    return re.sub(r"\s+v\d+$", "", _app.activeDocument.name)


def do_export(silent=True, outdir=None, label="", force_history=False):
    design = adsk.fusion.Design.cast(_app.activeProduct)
    if not design:
        if not silent:
            _ui.messageBox("No active design to export.", "MassTrack")
        return None

    if not _doc_is_saved():
        if not silent:
            _ui.messageBox(
                "Save the Fusion file first.\nMassTrack names its "
                "output folder after the document, so it needs a saved name.",
                "MassTrack")
        return None

    if outdir is None:
        outdir = _load_outdir()
    if not outdir or not os.path.isdir(outdir):
        if silent and not force_history:
            return None                        # save-triggered but not set up yet
        outdir = _ensure_outdir()
        if not outdir:
            return None

    ts = datetime.datetime.now().isoformat(timespec="seconds")
    docname = _doc_name()
    root = design.rootComponent
    doc_dir = _doc_dir(outdir, docname)
    doc_url = _doc_link_url(doc_dir)
    marked = list(_marked_items(design))
    rows, errors = [], []
    for key, kind, ent in marked:
        try:
            row = _props_row(key, kind, ent, design, ts, docname)
            row.append(_entity_link_url(ent, kind, design, doc_url))
            rows.append(row)
        except Exception:
            errors.append("%s (%s): %s" % (
                key, kind, traceback.format_exc(limit=1)))

    rows.sort(key=lambda r: (r[16].lower(), r[3].lower()))   # subassembly, name

    marked_comps = {}          # entityToken -> Component
    for _, kind, ent in marked:
        if kind == "component":
            try:
                marked_comps[ent.entityToken] = ent
            except Exception:
                pass
    marked_tokens = set(marked_comps)
    covered = set(marked_tokens)
    for comp in marked_comps.values():
        covered |= _descendant_tokens(comp, root)
    conflicts = []
    for token, comp in marked_comps.items():
        for dt in _descendant_tokens(comp, root) & marked_tokens:
            conflicts.append("%s contains marked %s"
                             % (comp.name, marked_comps[dt].name))
    for _, kind, ent in marked:
        if kind == "body":
            try:
                pc = ent.parentComponent
                if pc.entityToken in covered:
                    conflicts.append("%s contains marked body %s"
                                     % (pc.name, ent.name))
            except Exception:
                pass

    path = _csv_path(outdir, docname)
    tmp = path + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        w.writerows(rows)
    os.replace(tmp, path)

    # a double-count makes the total wrong, so don't record it into the curve
    if conflicts:
        total_eff = sum(float(r[11]) for r in rows)
        appended = False
    else:
        appended, total_eff = _update_history(
            outdir, docname, ts, rows, label, force_history)

    _log("export %s: %d rows, %d err, %d nesting, hist=%s, total=%.1fg, doc=%s"
         % (os.path.basename(path), len(rows), len(errors), len(conflicts),
            appended, total_eff, docname))
    for c in conflicts:
        _log("  DOUBLE-COUNT " + c)

    if not silent:
        msg = ("Exported %d marked part(s) from '%s'.\nTotal: %.1f g (%.3f kg)"
               "\nFile: %s" % (len(rows), docname, total_eff, total_eff / 1000.0,
                               os.path.basename(path)))
        if appended:
            msg += "\n\nHistory point added%s." % (
                ((": " + label) if label else ""))
        if errors:
            msg += "\n\n%d part(s) failed, see the MassTrack log." % len(errors)
        if conflicts:
            msg += ("\n\nDouble-count: %d marked component(s) sit inside "
                    "another marked component (mass counted twice). Unmark the "
                    "parent OR the child:\n  " % len(conflicts)) \
                + "\n  ".join(conflicts[:8])
        _ui.messageBox(msg, "MassTrack")
    return path


class _CmdCreated(adsk.core.CommandCreatedEventHandler):
    def __init__(self, cmd_id):
        super().__init__()
        self.cmd_id = cmd_id

    def notify(self, args):
        try:
            cmd = args.command
            if self.cmd_id in ("mtxMark", "mtxUnmark", "mtxSetKnown",
                               "mtxClearKnown"):
                pick = cmd.commandInputs.addDropDownCommandInput(
                    "pick", "Select",
                    adsk.core.DropDownStyles.TextListDropDownStyle)
                pick.listItems.add("Bodies only", True)
                pick.listItems.add("Components & bodies", False)
                pick.listItems.add("Components only", False)
                sel = cmd.commandInputs.addSelectionInput(
                    "sel", "Selection",
                    "Bodies (default), from the browser or canvas")
                sel.addSelectionFilter("Bodies")
                sel.setSelectionLimits(1, 0)
                on_input = _InputChanged()
                cmd.inputChanged.add(on_input)
                _handlers.append(on_input)
            if self.cmd_id == "mtxSetKnown":
                cmd.commandInputs.addStringValueInput(
                    "grams", "Known mass per unit (g)", "")
                cmd.commandInputs.addStringValueInput(
                    "note", "Note (spec / measured source)", "")
            if self.cmd_id == "mtxSnapshot":
                cmd.commandInputs.addStringValueInput(
                    "label", "History-point label", "")
            if self.cmd_id == "mtxSetHub":
                cmd.commandInputs.addStringValueInput(
                    "hub", "Fusion Team base URL",
                    _load_setting("hub_base") or "https://xxxx.autodesk360.com")
            on_exec = _CmdExecute(self.cmd_id)
            cmd.execute.add(on_exec)
            _handlers.append(on_exec)
        except Exception:
            _ui.messageBox("MassTrack error:\n" + traceback.format_exc())


class _InputChanged(adsk.core.InputChangedEventHandler):
    """Retarget the selection filter when the Components/Bodies picker changes."""
    def notify(self, args):
        try:
            if args.input.id != "pick":
                return
            sel = args.inputs.itemById("sel")
            if not sel:
                return
            choice = args.input.selectedItem.name
            sel.clearSelection()
            sel.clearSelectionFilter()
            if choice == "Components only":
                sel.addSelectionFilter("Occurrences")
            elif choice == "Bodies only":
                sel.addSelectionFilter("Bodies")
            else:
                sel.addSelectionFilter("Occurrences")
                sel.addSelectionFilter("Bodies")
        except Exception:
            _log("input-changed failed:\n" + traceback.format_exc())


class _CmdExecute(adsk.core.CommandEventHandler):
    def __init__(self, cmd_id):
        super().__init__()
        self.cmd_id = cmd_id

    def notify(self, args):
        try:
            design = adsk.fusion.Design.cast(_app.activeProduct)

            if self.cmd_id == "mtxGenerate":
                path = do_export(silent=False)
                if path:
                    _build_workbook(os.path.dirname(path))

            elif self.cmd_id == "mtxSnapshot":
                label = args.command.commandInputs.itemById("label").value.strip()
                path = do_export(silent=False, label=label, force_history=True)
                if path:
                    _refresh_workbook(os.path.dirname(path))

            elif self.cmd_id == "mtxDiagram":
                path = do_export(silent=False)
                if not path:
                    return
                doc_dir = os.path.dirname(path)
                png = os.path.join(doc_dir, "diagram.png")
                if not _capture_view(png):
                    _ui.messageBox("Could not capture the view.", "MassTrack")
                    return
                # _build_workbook embeds the diagram from the PNG it just wrote,
                # and re-embeds it on every later rebuild so it never vanishes
                out = _refresh_workbook(doc_dir)
                if not out:
                    _ui.messageBox(
                        "Could not build the workbook. If it is open in Excel, "
                        "close it and try again. Otherwise see the MassTrack log.",
                        "MassTrack")

            elif self.cmd_id == "mtxSetFolder":
                folder = _prompt_outdir()
                if folder:
                    _ui.messageBox("Output folder set to:\n" + folder,
                                   "MassTrack")

            elif self.cmd_id == "mtxSetHub":
                url = args.command.commandInputs.itemById("hub").value.strip()
                if url and "autodesk360.com" in url:
                    _save_setting("hub_base", url.rstrip("/"))
                    _ui.messageBox("Hub URL saved. Rows will now link to the "
                                   "file (login still required to open).",
                                   "MassTrack")
                elif url:
                    _ui.messageBox("That doesn't look like a Fusion Team URL "
                                   "(expected …autodesk360.com).", "MassTrack")

            elif self.cmd_id in ("mtxMark", "mtxUnmark", "mtxSetKnown",
                                 "mtxClearKnown"):
                inputs = args.command.commandInputs
                sel_input = inputs.itemById("sel")
                ents = [sel_input.selection(i).entity
                        for i in range(sel_input.selectionCount)]
                if self.cmd_id == "mtxMark":
                    verb = "Marked %d new item(s) (CAD mass)." % _mark(ents)
                elif self.cmd_id == "mtxUnmark":
                    verb = "Unmarked %d item(s)." % _unmark(ents)
                elif self.cmd_id == "mtxClearKnown":
                    _clear_known_mass(ents)
                    verb = "Reverted selection to CAD mass."
                else:                            # massSetKnown
                    grams = inputs.itemById("grams").value.strip().replace(",", ".")
                    note = inputs.itemById("note").value.strip()
                    try:
                        g = float(grams)
                        ok = math.isfinite(g) and g >= 0
                    except ValueError:
                        ok = False
                    if not ok:
                        _ui.messageBox("Enter a mass in grams, 0 or more.",
                                       "MassTrack")
                        return
                    n = _set_known_mass(ents, grams, note)
                    verb = ("Set known mass %s g on %d item(s)%s."
                            % (grams, n, ((": " + note) if note else "")))
                do_export(silent=True)
                _ui.messageBox(verb, "MassTrack")

            elif self.cmd_id == "mtxShow":
                if not design:
                    return
                ts = datetime.datetime.now().isoformat(timespec="seconds")
                docname = _doc_name()
                lines, total = [], 0.0
                for key, kind, ent in _marked_items(design):
                    try:
                        r = _props_row(key, kind, ent, design, ts, docname)
                        eff = float(r[11])          # mass_g_total
                        total += eff
                        tag = r[9]                  # cad / manual
                        extra = ("  (cad %.0f g)" % float(r[8])
                                 if tag == "manual" else "")
                        qty = "" if r[6] == "1" else "  x%s" % r[6]
                        note = ("  %s" % r[14]) if r[14] else ""
                        warn = ("  %s" % r[13]) if r[13] else ""
                        lines.append("%s%s\n    %.1f g  [%s]%s%s%s"
                                     % (r[3], qty, eff, tag, extra, note, warn))
                    except Exception:
                        lines.append("%s  [error reading]" % key)
                _ui.messageBox(
                    "Marked in '%s': %d part(s), total %.1f g (%.3f kg)\n\n%s"
                    % (docname, len(lines), total, total / 1000.0,
                       "\n".join(lines) or "(nothing marked)"), "MassTrack")

            elif self.cmd_id == "mtxHighlight":
                if not design:
                    return
                sels = _ui.activeSelections
                sels.clear()
                root = design.rootComponent
                n = 0
                for key, kind, ent in _marked_items(design):
                    try:
                        if kind == "body":
                            # attributes live on the NATIVE body; each occurrence
                            # needs its own proxy or bodies inside sub-occurrences
                            # never highlight in the assembly context.
                            occs = root.allOccurrencesByComponent(
                                ent.parentComponent)
                            if occs and occs.count > 0:
                                added = False
                                for i in range(occs.count):
                                    try:
                                        sels.add(ent.createForAssemblyContext(
                                            occs.item(i)))
                                        added = True
                                    except Exception:
                                        _log("highlight: no proxy for body '%s' "
                                             "in %s" % (ent.name,
                                                        occs.item(i).fullPathName))
                                if added:
                                    n += 1
                            else:                # root-level body: native is valid
                                sels.add(ent)
                                n += 1
                        else:                    # component: select its instances
                            occs = root.allOccurrencesByComponent(ent)
                            if occs and occs.count > 0:
                                added = False
                                for i in range(occs.count):
                                    try:
                                        sels.add(occs.item(i))
                                        added = True
                                    except Exception:
                                        _log("highlight: cannot select occurrence "
                                             "%s" % occs.item(i).fullPathName)
                                if added:
                                    n += 1
                            else:                # root/no occurrence: its bodies
                                for b in ent.bRepBodies:
                                    sels.add(b)
                                n += 1
                    except Exception:
                        pass
                if n == 0:
                    _ui.messageBox("Nothing is marked in this document.",
                                   "MassTrack")

            elif self.cmd_id == "mtxOpenXlsx":
                path = do_export(silent=False)
                if not path:
                    return
                doc_dir = os.path.dirname(path)
                wb = _refresh_workbook(doc_dir)
                if not wb:
                    xlsx = os.path.join(doc_dir,
                                        os.path.basename(doc_dir) + ".xlsx")
                    if os.path.isfile(xlsx):
                        _ui.messageBox(
                            "The workbook is open in Excel, so it could not "
                            "update. Close it and click Open in Excel again.",
                            "MassTrack")
                        _open_externally(xlsx)
                    else:
                        _ui.messageBox("Could not build the workbook. "
                                       "See the MassTrack log.", "MassTrack")
        except Exception:
            _ui.messageBox("MassTrack error:\n" + traceback.format_exc())


class _DocSaved(adsk.core.DocumentEventHandler):
    def notify(self, args):
        try:
            do_export(silent=True)              # no-op until an output folder is set
        except Exception:
            _log("documentSaved export failed:\n" + traceback.format_exc())


RES_DIR = os.path.join(ADDIN_DIR, "resources")
PANEL_ID = "MassTrackPanelX"
# promoted onto the ribbon, rest in overflow
PROMOTED = {"mtxGenerate", "mtxOpenXlsx", "mtxMark", "mtxSetKnown"}


def _make_panel():
    """MassTrack ribbon panel, falling back to ADD-INS if the Solid tab is missing.
    Reuses the panel if it already exists so Fusion keeps the user's ribbon
    arrangement (which buttons are promoted, their order) across reloads."""
    try:
        ws = _ui.workspaces.itemById("FusionSolidEnvironment")
        tab = ws.toolbarTabs.itemById("SolidTab") or ws.toolbarTabs.item(0)
        panel = tab.toolbarPanels.itemById(PANEL_ID)
        if not panel:
            panel = tab.toolbarPanels.add(PANEL_ID, "MassTrack")
        return panel
    except Exception:
        _log("panel create failed, using ADD-INS panel:\n"
             + traceback.format_exc())
        return _ui.allToolbarPanels.itemById("SolidScriptsAddinsPanel")


def run(context):
    global _app, _ui
    try:
        _app = adsk.core.Application.get()
        _ui = _app.userInterface
        # remove the previous version's panel if it lingers from before the
        # command IDs were rotated, so there is only one MassTrack panel
        try:
            _ws = _ui.workspaces.itemById("FusionSolidEnvironment")
            _tab = _ws.toolbarTabs.itemById("SolidTab") or _ws.toolbarTabs.item(0)
            _old = _tab.toolbarPanels.itemById("MassTrackPanel")
            if _old:
                _old.deleteMe()
        except Exception:
            pass
        panel = _make_panel()
        for cmd_id, cmd_name, tooltip in CMDS:
            # delete + recreate so every Stop->Run picks up fresh code AND fresh
            # icons (a reused definition would keep its old icon and handlers)
            existing = _ui.commandDefinitions.itemById(cmd_id)
            if existing:
                existing.deleteMe()
            icon = os.path.join(RES_DIR, cmd_id)
            if os.path.isdir(icon):
                cmd_def = _ui.commandDefinitions.addButtonDefinition(
                    cmd_id, cmd_name, tooltip, icon)
            else:
                cmd_def = _ui.commandDefinitions.addButtonDefinition(
                    cmd_id, cmd_name, tooltip)
            on_created = _CmdCreated(cmd_id)
            cmd_def.commandCreated.add(on_created)
            _handlers.append(on_created)
            _cmd_defs.append(cmd_def)
            if panel:
                ctl = panel.controls.itemById(cmd_id)
                if ctl:
                    ctl.deleteMe()
                ctl = panel.controls.addCommand(cmd_def)
                ctl.isPromoted = cmd_id in PROMOTED
                ctl.isPromotedByDefault = cmd_id in PROMOTED
                _controls.append(ctl)

        on_saved = _DocSaved()
        _app.documentSaved.add(on_saved)
        _handlers.append(on_saved)
        _log("add-in started [build=mtx-icons2]")
    except Exception:
        if _ui:
            _ui.messageBox("MassTrack failed to start:\n"
                           + traceback.format_exc())


def stop(context):
    try:
        for ctl in _controls:
            try:
                ctl.deleteMe()
            except Exception:
                pass
        for cd in _cmd_defs:
            try:
                cd.deleteMe()
            except Exception:
                pass
        _controls.clear()
        _cmd_defs.clear()
        _handlers.clear()
        _log("add-in stopped")
    except Exception:
        pass
